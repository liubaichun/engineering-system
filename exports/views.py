from rest_framework.views import APIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime


def style_header(ws, headers):
    """设置表头样式"""
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style='thin', color='E9ECEF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def style_data_rows(ws, row_count, col_count):
    """设置数据行样式"""
    thin = Side(style='thin', color='E9ECEF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill(start_color="F8F9FF", end_color="F8F9FF", fill_type="solid")
    for row in range(2, row_count + 2):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='center', horizontal='left')
            if row % 2 == 0:
                cell.fill = alt_fill


def auto_width(ws):
    """自动列宽"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def make_response(wb, filename):
    """生成 Excel 下载响应"""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    today = datetime.now().strftime('%Y%m%d')
    safe_name = filename.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{safe_name}_{today}.xlsx"'
    return response


# ────────────────────────────────────────────────────────
#  导出视图
# ────────────────────────────────────────────────────────

class ExportProjectsView(APIView):
    """导出所有项目为Excel"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from projects.models import Project
        projects = Project.objects.select_related('manager').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "项目清单"

        headers = ['序号', '项目名称', '客户', '供应商', '项目状态', '预算金额',
                   '开始日期', '结束日期', '项目经理', '项目描述', '创建时间']
        ws.append(headers)
        style_header(ws, headers)

        STATUS_MAP = dict(Project.STATUS_CHOICES)
        for idx, p in enumerate(projects, 1):
            ws.append([
                idx,
                p.name,
                p.client.name if p.client else '',
                p.supplier.name if p.supplier else '',
                STATUS_MAP.get(p.status, p.status),
                str(p.budget) if p.budget else '',
                str(p.start_date) if p.start_date else '',
                str(p.end_date) if p.end_date else '',
                p.manager.username if p.manager else '',
                p.description or '',
                p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else '',
            ])

        style_data_rows(ws, len(projects), len(headers))
        auto_width(ws)
        return make_response(wb, '项目清单')


class ExportTasksView(APIView):
    """导出所有任务为Excel"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from tasks.models import Task
        tasks = Task.objects.select_related('project', 'manager').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "任务清单"

        headers = ['序号', '任务名称', '所属项目', '负责人', '任务状态', '进度%',
                   '开始日期', '结束日期', '任务描述', '创建时间']
        ws.append(headers)
        style_header(ws, headers)

        STATUS_MAP = dict(Task.STATUS_CHOICES)
        for idx, t in enumerate(tasks, 1):
            ws.append([
                idx,
                t.name,
                t.project.name if t.project else '',
                t.manager.username if t.manager else '',
                STATUS_MAP.get(t.status, t.status),
                t.progress,
                str(t.start_date) if t.start_date else '',
                str(t.end_date) if t.end_date else '',
                t.description or '',
                t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
            ])

        style_data_rows(ws, len(tasks), len(headers))
        auto_width(ws)
        return make_response(wb, '任务清单')


class ExportMaterialsView(APIView):
    """导出物料清单为Excel"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from inventory.models import MaterialNew
        materials = MaterialNew.objects.select_related('supplier').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "物料清单"

        headers = ['序号', '物料名称', '规格型号', '单位', '当前库存', '预警阈值', '供应商', '创建时间']
        ws.append(headers)
        style_header(ws, headers)

        UNIT_MAP = dict(MaterialNew.UNIT_CHOICES)
        for idx, m in enumerate(materials, 1):
            ws.append([
                idx,
                m.name,
                m.specification or '',
                UNIT_MAP.get(m.unit, m.unit),
                str(m.stock) if m.stock else '0',
                str(m.alert_threshold) if m.alert_threshold else '0',
                m.supplier.name if m.supplier else '',
                m.created_at.strftime('%Y-%m-%d %H:%M') if m.created_at else '',
            ])

        style_data_rows(ws, len(materials), len(headers))
        auto_width(ws)
        return make_response(wb, '物料清单')


class ExportEquipmentView(APIView):
    """导出设备清单为Excel"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from inventory.models import EquipmentNew
        equipment = EquipmentNew.objects.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "设备清单"

        headers = ['序号', '设备名称', '规格型号', '设备型号', '设备状态', '存放地点', '创建时间']
        ws.append(headers)
        style_header(ws, headers)

        STATUS_MAP = dict(EquipmentNew.STATUS_CHOICES)
        for idx, e in enumerate(equipment, 1):
            ws.append([
                idx,
                e.name,
                e.specification or '',
                e.model or '',
                STATUS_MAP.get(e.status, e.status),
                e.location or '',
                e.created_at.strftime('%Y-%m-%d %H:%M') if e.created_at else '',
            ])

        style_data_rows(ws, len(equipment), len(headers))
        auto_width(ws)
        return make_response(wb, '设备清单')


# ────────────────────────────────────────────────────────
#  导入视图
# ────────────────────────────────────────────────────────

class ImportProjectsView(APIView):
    """批量导入项目"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from projects.models import Project
        from projects.serializers import ProjectSerializer
        import openpyxl
        from io import BytesIO

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': '请上传 Excel 文件'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
        except Exception:
            return Response({'detail': '无法读取 Excel 文件，请确认格式正确'}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        # 验证表头
        expected_headers = ['项目名称', '客户', '供应商', '项目状态', '预算金额', '开始日期', '结束日期', '项目描述']
        actual_headers = [cell.value for cell in ws[1]]
        if actual_headers[:len(expected_headers)] != expected_headers:
            return Response({
                'detail': f'表头不匹配，期望: {expected_headers}，实际: {actual_headers[:len(expected_headers)]}'
            }, status=status.HTTP_400_BAD_REQUEST)

        STATUS_CHOICES = [c[0] for c in Project.STATUS_CHOICES]
        created = []
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = row[0]
            if not name:
                errors.append(f'第{row_idx}行：项目名称不能为空')
                continue

            status_val = row[3]
            if status_val and status_val not in STATUS_CHOICES:
                # 尝试中文状态名匹配
                status_map = {c[1]: c[0] for c in Project.STATUS_CHOICES}
                status_val = status_map.get(status_val, 'preparing')

            data = {
                'name': str(name).strip(),
                'client': str(row[1] or '').strip(),
                'supplier': str(row[2] or '').strip(),
                'status': status_val or 'preparing',
                'budget': row[4] or 0,
                'start_date': str(row[5])[:10] if row[5] else None,
                'end_date': str(row[6])[:10] if row[6] else None,
                'description': str(row[7] or '').strip(),
            }

            serializer = ProjectSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                created.append(name)
            else:
                errors.append(f'第{row_idx}行「{name}」：{serializer.errors}')

        return Response({
            'created': len(created),
            'errors': errors,
            'detail': f'成功导入 {len(created)} 个项目，{len(errors)} 个错误'
        }, status=status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST)
