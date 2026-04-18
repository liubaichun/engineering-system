from rest_framework.decorators import action
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import Income, Expense, InvoiceNew, FinancialRecord, Invoice, Company, Salary, WageRecord
from .serializers import (
    IncomeSerializer, ExpenseSerializer,
    InvoiceNewSerializer, FinancialRecordSerializer, InvoiceSerializer,
    CompanySerializer, SalarySerializer, WageRecordSerializer
)
from .permissions import IsFinanceOnly
from operation_logs.models import OperationLog


class IncomeViewSet(viewsets.ModelViewSet):
    """收入视图集"""
    queryset = Income.objects.select_related('project', 'customer', 'operator').all()
    serializer_class = IncomeSerializer
    permission_classes = [IsAuthenticated, IsFinanceOnly]

    def get_client_ip(self):
        """获取客户端IP"""
        x_forwarded_for = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = self.request.META.get('REMOTE_ADDR')
        return ip

    def perform_create(self, serializer):
        # Auto-set operator from request user
        obj = serializer.save(operator=self.request.user)
        desc = f"创建了收入记录：金额 {obj.amount}"
        if obj.project:
            desc += f"，项目：{obj.project.name}"
        OperationLog.objects.create(
            user=self.request.user,
            action='create',
            model_name='Income',
            object_id=obj.id,
            description=desc,
            ip_address=self.get_client_ip()
        )

    def perform_update(self, serializer):
        obj = serializer.save()
        desc = f"更新了收入记录：金额 {obj.amount}"
        if obj.project:
            desc += f"，项目：{obj.project.name}"
        OperationLog.objects.create(
            user=self.request.user,
            action='update',
            model_name='Income',
            object_id=obj.id,
            description=desc,
            ip_address=self.get_client_ip()
        )

    def perform_destroy(self, instance):
        income_id = instance.id
        amount = instance.amount
        instance.delete()
        OperationLog.objects.create(
            user=self.request.user,
            action='delete',
            model_name='Income',
            object_id=income_id,
            description=f"删除了收入记录：金额 {amount}",
            ip_address=self.get_client_ip()
        )


class ExpenseViewSet(viewsets.ModelViewSet):
    """支出视图集"""
    queryset = Expense.objects.select_related('project', 'supplier', 'operator').all()
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated, IsFinanceOnly]

    def get_client_ip(self):
        """获取客户端IP"""
        x_forwarded_for = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = self.request.META.get('REMOTE_ADDR')
        return ip

    def perform_create(self, serializer):
        # Auto-set operator from request user
        obj = serializer.save(operator=self.request.user)
        desc = f"创建了支出记录：金额 {obj.amount}"
        if obj.project:
            desc += f"，项目：{obj.project.name}"
        OperationLog.objects.create(
            user=self.request.user,
            action='create',
            model_name='Expense',
            object_id=obj.id,
            description=desc,
            ip_address=self.get_client_ip()
        )

    def perform_update(self, serializer):
        obj = serializer.save()
        desc = f"更新了支出记录：金额 {obj.amount}"
        if obj.project:
            desc += f"，项目：{obj.project.name}"
        OperationLog.objects.create(
            user=self.request.user,
            action='update',
            model_name='Expense',
            object_id=obj.id,
            description=desc,
            ip_address=self.get_client_ip()
        )

    def perform_destroy(self, instance):
        expense_id = instance.id
        amount = instance.amount
        instance.delete()
        OperationLog.objects.create(
            user=self.request.user,
            action='delete',
            model_name='Expense',
            object_id=expense_id,
            description=f"删除了支出记录：金额 {amount}",
            ip_address=self.get_client_ip()
        )


class InvoiceNewViewSet(viewsets.ModelViewSet):
    """发票视图集（新）"""
    queryset = InvoiceNew.objects.all()
    serializer_class = InvoiceNewSerializer
    permission_classes = [IsAuthenticated, IsFinanceOnly]


class FinancialRecordViewSet(viewsets.ModelViewSet):
    """财务记录视图集（兼容）"""
    queryset = FinancialRecord.objects.all()
    serializer_class = FinancialRecordSerializer
    permission_classes = [IsAuthenticated, IsFinanceOnly]


class InvoiceViewSet(viewsets.ModelViewSet):
    """发票视图集（兼容）"""
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer
    permission_classes = [IsAuthenticated, IsFinanceOnly]


class CompanyViewSet(viewsets.ModelViewSet):
    """公司视图集"""
    queryset = Company.objects.all()
    serializer_class = CompanySerializer
    permission_classes = [IsAuthenticated, IsFinanceOnly]

    def destroy(self, request, *args, **kwargs):
        """重写destroy方法，处理ProtectedError"""
        from django.db.models import ProtectedError
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
            return Response(status=204)
        except ProtectedError:
            # 获取关联的模型信息
            related_models = []
            for rel in instance._meta.related_objects:
                if rel.related_model:
                    related_models.append(rel.related_model._meta.verbose_name)
            return Response(
                {'error': f'无法删除该公司，存在关联的{related_models}记录，请先删除关联记录'},
                status=400
            )


class SalaryViewSet(viewsets.ModelViewSet):
    """工资单视图集"""
    queryset = Salary.objects.select_related('company', 'approver').all()
    serializer_class = SalarySerializer
    permission_classes = [IsAuthenticated, IsFinanceOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        company_id = self.request.query_params.get('company_id')
        month = self.request.query_params.get('month')
        if company_id:
            queryset = queryset.filter(company_id=company_id)
        if month:
            queryset = queryset.filter(salary_month=month)
        return queryset

    def get_client_ip(self):
        x_forwarded_for = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = self.request.META.get('REMOTE_ADDR')
        return ip

    def perform_create(self, serializer):
        obj = serializer.save()
        OperationLog.objects.create(
            user=self.request.user,
            action='create',
            model_name='Salary',
            object_id=obj.id,
            description=f"创建了工资单：{obj.employee_id} {obj.salary_month}",
            ip_address=self.get_client_ip()
        )

    def perform_update(self, serializer):
        obj = serializer.save()
        OperationLog.objects.create(
            user=self.request.user,
            action='update',
            model_name='Salary',
            object_id=obj.id,
            description=f"更新了工资单：{obj.employee_id} {obj.salary_month}",
            ip_address=self.get_client_ip()
        )

    @action(detail=False, methods=['get'])
    def export(self, request):
        导出工资单Excel
        company_id = request.query_params.get('company_id')
        month = request.query_params.get('month')

        if not company_id:
            return Response({'error': '缺少 company_id 参数'}, status=400)

        queryset = Salary.objects.select_related('company').filter(company_id=company_id)
        if month:
            queryset = queryset.filter(month=month)

        if not queryset.exists():
            return Response({'error': '没有找到工资单数据'}, status=404)

        company_name = queryset.first().company.name

        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = f'{month or "全部"}工资单'

        headers = [
            '序号', '员工姓名', '工号', '职位', '银行卡号', '月份',
            '基本工资', '加班工资', '奖金', '社保扣款', '公积金扣款',
            '员工借款', '其他扣款', '其他加款', '个税', '实发工资', '状态', '备注'
        ]
        header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        status_map = dict(Salary.STATUS_CHOICES)
        for row_idx, salary in enumerate(queryset, 2):
            row_data = [
                row_idx - 1,
                salary.employee_name,
                salary.employee_no,
                salary.position,
                salary.bank_account,
                salary.month,
                float(salary.base_salary),
                float(salary.overtime_salary),
                float(salary.bonus),
                float(salary.social_insurance),
                float(salary.housing_fund),
                float(salary.employee_loan or 0),
                float(salary.other_deduction or 0),
                float(salary.other_addition or 0),
                float(salary.individual_income_tax),
                float(salary.net_salary),
                status_map.get(salary.status, salary.status),
                salary.remark or '',
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if col >= 7:  # 数值列右对齐
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='center')

        # 设置金额列格式
        money_cols = range(7, 16)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=15):
            for cell in row:
                cell.number_format = '#,##0.00'

        # 自动列宽
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'{company_name}_{month or "全部"}_工资单.xlsx'
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
        return response


class WageRecordViewSet(viewsets.ModelViewSet):
    """工资单记录视图集"""
    queryset = WageRecord.objects.select_related('company', 'approver').all()
    serializer_class = WageRecordSerializer
    permission_classes = [IsAuthenticated, IsFinanceOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        company_id = self.request.query_params.get('company_id')
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        status = self.request.query_params.get('status')

        if company_id:
            queryset = queryset.filter(company_id=company_id)
        if year:
            queryset = queryset.filter(year=year)
        if month:
            queryset = queryset.filter(month=month)
        if status:
            queryset = queryset.filter(status=status)
        return queryset

    def get_client_ip(self):
        x_forwarded_for = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = self.request.META.get('REMOTE_ADDR')
        return ip

    def perform_create(self, serializer):
        obj = serializer.save()
        OperationLog.objects.create(
            user=self.request.user,
            action='create',
            model_name='WageRecord',
            object_id=obj.id,
            description=f"创建了工资单：{obj.employee_name} {obj.year}-{obj.month:02d}",
            ip_address=self.get_client_ip()
        )

    def perform_update(self, serializer):
        obj = serializer.save()
        OperationLog.objects.create(
            user=self.request.user,
            action='update',
            model_name='WageRecord',
            object_id=obj.id,
            description=f"更新了工资单：{obj.employee_name} {obj.year}-{obj.month:02d}",
            ip_address=self.get_client_ip()
        )

    @action(detail=False, methods=['get'])
    def export(self, request):
        """导出工资单Excel"""
        company_id = request.query_params.get('company_id')
        year = request.query_params.get('year')
        month = request.query_params.get('month')

        if not company_id:
            return Response({'error': '缺少 company_id 参数'}, status=400)

        queryset = WageRecord.objects.select_related('company').filter(company_id=company_id)
        if year:
            queryset = queryset.filter(year=year)
        if month:
            queryset = queryset.filter(month=month)

        if not queryset.exists():
            return Response({'error': '没有找到工资单数据'}, status=404)

        company_name = queryset.first().company.name
        period = f"{year or '全部'}-{month or '全部'}"

        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = f'{period}工资单'

        headers = [
            '序号', '员工姓名', '部门', '职位', '银行卡号', '月份',
            '基本工资', '加班费', '奖金', '社保扣款', '公积金扣款',
            '请假扣款', '其他扣款', '个税前工资', '个税', '实发工资', '状态', '备注'
        ]
        header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        status_map = dict(WageRecord.STATUS_CHOICES)
        for row_idx, wage in enumerate(queryset, 2):
            row_data = [
                row_idx - 1,
                wage.employee_name,
                wage.department,
                wage.position,
                wage.bank_card,
                f"{wage.year}-{wage.month:02d}",
                float(wage.base_salary),
                float(wage.overtime_pay or 0),
                float(wage.bonus or 0),
                float(wage.social_insurance or 0),
                float(wage.housing_fund or 0),
                float(wage.leave_deduction or 0),
                float(wage.other_deductions or 0),
                float(wage.gross_salary),
                float(wage.tax or 0),
                float(wage.net_salary),
                status_map.get(wage.status, wage.status),
                wage.remarks or '',
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if col >= 7:  # 数值列右对齐
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='center')

        # 设置金额列格式
        money_cols = range(7, 16)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=15):
            for cell in row:
                cell.number_format = '#,##0.00'

        # 自动列宽
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'{company_name}_{period}_工资单.xlsx'
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
        return response


class WageReportViewSet(viewsets.ViewSet):
    """工资报表视图集"""
    permission_classes = [IsAuthenticated, IsFinanceOnly]

    def list(self, request):
        """获取工资报表数据"""
        from django.db.models import Sum, Count
        from finance.models import WageRecord

        company_id = request.query_params.get('company_id')
        year = request.query_params.get('year')
        report_type = request.query_params.get('type', 'monthly')  # monthly, quarterly, yearly

        if not company_id:
            return Response({'error': '缺少 company_id 参数'}, status=400)

        queryset = WageRecord.objects.filter(company_id=company_id)
        if year:
            queryset = queryset.filter(year=year)

        if report_type == 'monthly':
            return self._monthly_report(queryset, year)
        elif report_type == 'quarterly':
            return self._quarterly_report(queryset, year)
        elif report_type == 'yearly':
            return self._yearly_report(queryset, year)
        else:
            return Response({'error': '无效的报表类型'}, status=400)

    def _monthly_report(self, queryset, year):
        """月度报表"""
        if not year:
            year = str(2026)

        months_data = []
        for m in range(1, 13):
            month_qs = queryset.filter(year=year, month=m)
            stats = month_qs.aggregate(
                employee_count=Count('id'),
                total_base=Sum('base_salary'),
                total_overtime=Sum('overtime_pay'),
                total_bonus=Sum('bonus'),
                total_gross=Sum('gross_salary'),
                total_social=Sum('social_insurance'),
                total_housing=Sum('housing_fund'),
                total_leave=Sum('leave_deduction'),
                total_other=Sum('other_deductions'),
                total_tax=Sum('tax'),
                total_net=Sum('net_salary'),
            )
            months_data.append({
                'year': int(year),
                'month': m,
                'month_display': f'{year}年{m}月',
                'employee_count': stats['employee_count'] or 0,
                'total_base': float(stats['total_base'] or 0),
                'total_overtime': float(stats['total_overtime'] or 0),
                'total_bonus': float(stats['total_bonus'] or 0),
                'total_gross': float(stats['total_gross'] or 0),
                'total_social': float(stats['total_social'] or 0),
                'total_housing': float(stats['total_housing'] or 0),
                'total_leave': float(stats['total_leave'] or 0),
                'total_other': float(stats['total_other'] or 0),
                'total_tax': float(stats['total_tax'] or 0),
                'total_net': float(stats['total_net'] or 0),
            })

        # 计算合计
        total_count = sum(m['employee_count'] for m in months_data)
        totals = {
            'total_base': sum(m['total_base'] for m in months_data),
            'total_overtime': sum(m['total_overtime'] for m in months_data),
            'total_bonus': sum(m['total_bonus'] for m in months_data),
            'total_gross': sum(m['total_gross'] for m in months_data),
            'total_social': sum(m['total_social'] for m in months_data),
            'total_housing': sum(m['total_housing'] for m in months_data),
            'total_leave': sum(m['total_leave'] for m in months_data),
            'total_other': sum(m['total_other'] for m in months_data),
            'total_tax': sum(m['total_tax'] for m in months_data),
            'total_net': sum(m['total_net'] for m in months_data),
        }

        return Response({
            'type': 'monthly',
            'year': year,
            'data': months_data,
            'summary': {
                'total_employee_count': total_count,
                **totals
            }
        })

    def _quarterly_report(self, queryset, year):
        """季度报表"""
        if not year:
            year = str(2026)

        quarters_data = []
        for q in range(1, 5):
            month_range = [(q - 1) * 3 + i for i in range(1, 4)]
            quarter_qs = queryset.filter(year=year, month__in=month_range)
            stats = quarter_qs.aggregate(
                employee_count=Count('id'),
                total_base=Sum('base_salary'),
                total_overtime=Sum('overtime_pay'),
                total_bonus=Sum('bonus'),
                total_gross=Sum('gross_salary'),
                total_social=Sum('social_insurance'),
                total_housing=Sum('housing_fund'),
                total_leave=Sum('leave_deduction'),
                total_other=Sum('other_deductions'),
                total_tax=Sum('tax'),
                total_net=Sum('net_salary'),
            )
            quarters_data.append({
                'year': int(year),
                'quarter': q,
                'quarter_display': f'{year}年第{q}季度',
                'months': month_range,
                'employee_count': stats['employee_count'] or 0,
                'total_base': float(stats['total_base'] or 0),
                'total_overtime': float(stats['total_overtime'] or 0),
                'total_bonus': float(stats['total_bonus'] or 0),
                'total_gross': float(stats['total_gross'] or 0),
                'total_social': float(stats['total_social'] or 0),
                'total_housing': float(stats['total_housing'] or 0),
                'total_leave': float(stats['total_leave'] or 0),
                'total_other': float(stats['total_other'] or 0),
                'total_tax': float(stats['total_tax'] or 0),
                'total_net': float(stats['total_net'] or 0),
            })

        totals = {
            'total_base': sum(m['total_base'] for m in quarters_data),
            'total_overtime': sum(m['total_overtime'] for m in quarters_data),
            'total_bonus': sum(m['total_bonus'] for m in quarters_data),
            'total_gross': sum(m['total_gross'] for m in quarters_data),
            'total_social': sum(m['total_social'] for m in quarters_data),
            'total_housing': sum(m['total_housing'] for m in quarters_data),
            'total_leave': sum(m['total_leave'] for m in quarters_data),
            'total_other': sum(m['total_other'] for m in quarters_data),
            'total_tax': sum(m['total_tax'] for m in quarters_data),
            'total_net': sum(m['total_net'] for m in quarters_data),
        }

        return Response({
            'type': 'quarterly',
            'year': year,
            'data': quarters_data,
            'summary': totals
        })

    def _yearly_report(self, queryset, year):
        """年度报表"""
        if not year:
            year = str(2026)

        stats = queryset.aggregate(
            employee_count=Count('id'),
            total_base=Sum('base_salary'),
            total_overtime=Sum('overtime_pay'),
            total_bonus=Sum('bonus'),
            total_gross=Sum('gross_salary'),
            total_social=Sum('social_insurance'),
            total_housing=Sum('housing_fund'),
            total_leave=Sum('leave_deduction'),
            total_other=Sum('other_deductions'),
            total_tax=Sum('tax'),
            total_net=Sum('net_salary'),
        )

        return Response({
            'type': 'yearly',
            'year': year,
            'data': [{
                'year': int(year),
                'employee_count': stats['employee_count'] or 0,
                'total_base': float(stats['total_base'] or 0),
                'total_overtime': float(stats['total_overtime'] or 0),
                'total_bonus': float(stats['total_bonus'] or 0),
                'total_gross': float(stats['total_gross'] or 0),
                'total_social': float(stats['total_social'] or 0),
                'total_housing': float(stats['total_housing'] or 0),
                'total_leave': float(stats['total_leave'] or 0),
                'total_other': float(stats['total_other'] or 0),
                'total_tax': float(stats['total_tax'] or 0),
                'total_net': float(stats['total_net'] or 0),
            }],
            'summary': {
                'employee_count': stats['employee_count'] or 0,
                'total_base': float(stats['total_base'] or 0),
                'total_overtime': float(stats['total_overtime'] or 0),
                'total_bonus': float(stats['total_bonus'] or 0),
                'total_gross': float(stats['total_gross'] or 0),
                'total_social': float(stats['total_social'] or 0),
                'total_housing': float(stats['total_housing'] or 0),
                'total_leave': float(stats['total_leave'] or 0),
                'total_other': float(stats['total_other'] or 0),
                'total_tax': float(stats['total_tax'] or 0),
                'total_net': float(stats['total_net'] or 0),
            }
        })


class MonthlyReportViewSet(viewsets.ViewSet):
    """月度报表视图集"""
    permission_classes = [IsAuthenticated, IsFinanceOnly]

    def list(self, request):
        """获取月度报表数据"""
        from django.db.models import Sum, Count
        from django.db.models.functions import TruncMonth
        from finance.models import Income, Expense

        year = request.query_params.get('year')
        if not year:
            from datetime import datetime
            year = str(datetime.now().year)

        # 收入汇总（按月）
        income_by_month = Income.objects.filter(
            date__startswith=year
        ).annotate(
            month=TruncMonth('date')
        ).values('month').annotate(
            total_amount=Sum('amount'),
            count=Count('id')
        ).order_by('month')

        # 支出汇总（按月）
        expense_by_month = Expense.objects.filter(
            date__startswith=year
        ).annotate(
            month=TruncMonth('date')
        ).values('month').annotate(
            total_amount=Sum('amount'),
            count=Count('id')
        ).order_by('month')

        # 构建月度数据
        months_data = {}
        for i in range(1, 13):
            month_key = f"{year}-{i:02d}"
            months_data[month_key] = {
                'month': month_key,
                'income': 0,
                'income_count': 0,
                'expense': 0,
                'expense_count': 0,
                'balance': 0
            }

        for item in income_by_month:
            month_key = item['month'].strftime('%Y-%m') if item['month'] else None
            if month_key and month_key in months_data:
                months_data[month_key]['income'] = float(item['total_amount'] or 0)
                months_data[month_key]['income_count'] = item['count']

        for item in expense_by_month:
            month_key = item['month'].strftime('%Y-%m') if item['month'] else None
            if month_key and month_key in months_data:
                months_data[month_key]['expense'] = float(item['total_amount'] or 0)
                months_data[month_key]['expense_count'] = item['count']

        # 计算月度结余
        result = []
        cumulative_balance = 0
        for month_key in sorted(months_data.keys()):
            data = months_data[month_key]
            data['balance'] = data['income'] - data['expense']
            cumulative_balance += data['balance']
            data['cumulative_balance'] = cumulative_balance
            result.append(data)

        return Response({
            'year': year,
            'monthly_data': result,
            'summary': {
                'total_income': sum(m['income'] for m in result),
                'total_expense': sum(m['expense'] for m in result),
                'total_balance': sum(m['balance'] for m in result)
            }
        })
